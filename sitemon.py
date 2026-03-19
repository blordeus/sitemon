"""
Site Performance Monitor
--------------------------
Audits websites for response time, HTTP status, redirects,
SSL certificate expiry, and page size.

Supports one-off audits and scheduled monitoring with history logging.

Usage:
    python sitemon.py --urls https://example.com https://another.com
    python sitemon.py --file sites.txt
    python sitemon.py --file sites.txt --output report.xlsx
    python sitemon.py --file sites.txt --schedule 30   (runs every 30 minutes)
    python sitemon.py --history

Setup:
    pip install requests openpyxl schedule
"""

import argparse
import json
import socket
import ssl
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

import requests
import schedule
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

# ─── CONFIG ───────────────────────────────────────────────────────────────────

HISTORY_FILE = Path.home() / ".sitemon_history.json"
TIMEOUT = 10  # seconds

# Status thresholds
RESPONSE_WARN_MS = 1500   # yellow warning
RESPONSE_FAIL_MS = 3000   # red alert
SSL_WARN_DAYS = 30        # warn if SSL expires within 30 days

# Colors
COLOR_GREEN  = "C6EFCE"
COLOR_YELLOW = "FFEB9C"
COLOR_RED    = "FFC7CE"
COLOR_HEADER = "365349"
COLOR_SUBHEADER = "435066"


# ─── AUDIT ────────────────────────────────────────────────────────────────────

def check_ssl(hostname: str) -> dict:
    """Check SSL certificate expiry for a hostname."""
    try:
        ctx = ssl.create_default_context()
        with ctx.wrap_socket(socket.socket(), server_hostname=hostname) as s:
            s.settimeout(TIMEOUT)
            s.connect((hostname, 443))
            cert = s.getpeercert()
            expires_str = cert["notAfter"]
            expires_dt = datetime.strptime(expires_str, "%b %d %H:%M:%S %Y %Z").replace(tzinfo=timezone.utc)
            days_left = (expires_dt - datetime.now(timezone.utc)).days
            return {
                "ssl_valid": True,
                "ssl_expiry": expires_dt.strftime("%Y-%m-%d"),
                "ssl_days_left": days_left,
                "ssl_status": "OK" if days_left > SSL_WARN_DAYS else ("WARNING" if days_left > 0 else "EXPIRED"),
            }
    except ssl.SSLError as e:
        return {"ssl_valid": False, "ssl_expiry": "N/A", "ssl_days_left": None, "ssl_status": f"SSL Error"}
    except Exception as e:
        return {"ssl_valid": False, "ssl_expiry": "N/A", "ssl_days_left": None, "ssl_status": "N/A"}


def audit_url(url: str) -> dict:
    """Run a full audit on a single URL."""
    if not url.startswith("http"):
        url = "https://" + url

    result = {
        "url": url,
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "status_code": None,
        "response_time_ms": None,
        "page_size_kb": None,
        "redirects": 0,
        "final_url": url,
        "redirect_chain": "",
        "ssl_valid": None,
        "ssl_expiry": None,
        "ssl_days_left": None,
        "ssl_status": None,
        "overall_status": "UNKNOWN",
        "error": None,
    }

    try:
        start = time.time()
        response = requests.get(
            url,
            timeout=TIMEOUT,
            headers={"User-Agent": "Mozilla/5.0 (compatible; SiteMonitor/1.0)"},
            allow_redirects=True,
        )
        elapsed_ms = round((time.time() - start) * 1000)

        # Redirect chain
        redirects = response.history
        chain = " → ".join([r.url for r in redirects] + [response.url]) if redirects else ""

        # Page size
        content_length = len(response.content)
        size_kb = round(content_length / 1024, 1)

        result.update({
            "status_code": response.status_code,
            "response_time_ms": elapsed_ms,
            "page_size_kb": size_kb,
            "redirects": len(redirects),
            "final_url": response.url,
            "redirect_chain": chain,
        })

        # SSL check (only for https)
        hostname = url.split("//")[-1].split("/")[0]
        if url.startswith("https"):
            ssl_info = check_ssl(hostname)
            result.update(ssl_info)

        # Overall status
        if response.status_code >= 400:
            result["overall_status"] = "ERROR"
        elif elapsed_ms >= RESPONSE_FAIL_MS:
            result["overall_status"] = "SLOW"
        elif elapsed_ms >= RESPONSE_WARN_MS or (result["ssl_days_left"] and result["ssl_days_left"] <= SSL_WARN_DAYS):
            result["overall_status"] = "WARNING"
        else:
            result["overall_status"] = "OK"

    except requests.exceptions.SSLError as e:
        result["error"] = "SSL Error"
        result["overall_status"] = "ERROR"
    except requests.exceptions.ConnectionError:
        result["error"] = "Connection failed"
        result["overall_status"] = "DOWN"
    except requests.exceptions.Timeout:
        result["error"] = f"Timed out after {TIMEOUT}s"
        result["overall_status"] = "DOWN"
    except Exception as e:
        result["error"] = str(e)
        result["overall_status"] = "ERROR"

    return result


def audit_all(urls: list) -> list:
    """Audit a list of URLs and return results."""
    results = []
    for i, url in enumerate(urls, 1):
        url = url.strip()
        if not url:
            continue
        print(f"  [{i}/{len(urls)}] Checking {url}...", end=" ", flush=True)
        result = audit_url(url)
        status_icon = {"OK": "✅", "WARNING": "⚠️", "SLOW": "🐢", "ERROR": "❌", "DOWN": "🔴"}.get(result["overall_status"], "❓")
        print(f"{status_icon} {result['overall_status']} ({result.get('response_time_ms', '—')}ms)")
        results.append(result)
    return results


# ─── INPUT ────────────────────────────────────────────────────────────────────

def load_urls_from_file(filepath: str) -> list:
    path = Path(filepath)
    if not path.exists():
        print(f"❌ File not found: {filepath}")
        sys.exit(1)
    lines = path.read_text(encoding="utf-8").splitlines()
    urls = [l.strip() for l in lines if l.strip() and not l.startswith("#")]
    print(f"  📋 Loaded {len(urls)} URLs from {filepath}")
    return urls


# ─── HISTORY ──────────────────────────────────────────────────────────────────

def load_history() -> list:
    if HISTORY_FILE.exists():
        with open(HISTORY_FILE, "r") as f:
            return json.load(f)
    return []


def save_to_history(results: list):
    history = load_history()
    history.extend(results)
    history = history[-500:]  # keep last 500 entries
    with open(HISTORY_FILE, "w") as f:
        json.dump(history, f, indent=2)
    print(f"  💾 Saved {len(results)} result(s) to history")


def show_history(n: int = 10):
    history = load_history()
    if not history:
        print("No history found.")
        return
    print(f"\n── Last {min(n, len(history))} checks ──\n")
    for entry in history[-n:]:
        icon = {"OK": "✅", "WARNING": "⚠️", "SLOW": "🐢", "ERROR": "❌", "DOWN": "🔴"}.get(entry.get("overall_status"), "❓")
        print(f"  {entry['timestamp']}  {icon} {entry['overall_status']:8}  {entry['url']}")
    print()


# ─── TERMINAL OUTPUT ──────────────────────────────────────────────────────────

def print_results(results: list):
    print("\n" + "=" * 70)
    print("  SITE AUDIT RESULTS")
    print("=" * 70)

    for r in results:
        print(f"\n  {r['url']}")
        print(f"    Status      : {r['status_code']} — {r['overall_status']}")
        print(f"    Response    : {r['response_time_ms']}ms")
        print(f"    Page Size   : {r['page_size_kb']} KB")
        print(f"    Redirects   : {r['redirects']}")
        if r["redirect_chain"]:
            print(f"    Chain       : {r['redirect_chain'][:80]}...")
        print(f"    SSL Expiry  : {r['ssl_expiry']} ({r['ssl_days_left']} days) — {r['ssl_status']}")
        if r["error"]:
            print(f"    ⚠️  Error    : {r['error']}")
    print()


# ─── EXCEL EXPORT ─────────────────────────────────────────────────────────────

COLUMNS = [
    ("URL", 40),
    ("Timestamp", 18),
    ("Overall Status", 14),
    ("HTTP Status", 12),
    ("Response (ms)", 14),
    ("Page Size (KB)", 14),
    ("Redirects", 10),
    ("Final URL", 35),
    ("SSL Expiry", 12),
    ("SSL Days Left", 13),
    ("SSL Status", 12),
    ("Error", 25),
]


def style_header(ws, fill_color=COLOR_HEADER):
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")


def apply_status_colors(ws, col_index: int, start_row: int = 2):
    """Color the Overall Status column based on value."""
    fills = {
        "OK":      PatternFill(start_color=COLOR_GREEN,  end_color=COLOR_GREEN,  fill_type="solid"),
        "WARNING": PatternFill(start_color=COLOR_YELLOW, end_color=COLOR_YELLOW, fill_type="solid"),
        "SLOW":    PatternFill(start_color=COLOR_YELLOW, end_color=COLOR_YELLOW, fill_type="solid"),
        "ERROR":   PatternFill(start_color=COLOR_RED,    end_color=COLOR_RED,    fill_type="solid"),
        "DOWN":    PatternFill(start_color=COLOR_RED,    end_color=COLOR_RED,    fill_type="solid"),
    }
    for row in ws.iter_rows(min_row=start_row, min_col=col_index, max_col=col_index):
        for cell in row:
            cell.fill = fills.get(cell.value, PatternFill())
            cell.font = Font(bold=True)


def results_to_rows(results: list) -> list:
    rows = []
    for r in results:
        rows.append([
            r.get("url"),
            r.get("timestamp"),
            r.get("overall_status"),
            r.get("status_code"),
            r.get("response_time_ms"),
            r.get("page_size_kb"),
            r.get("redirects"),
            r.get("final_url"),
            r.get("ssl_expiry"),
            r.get("ssl_days_left"),
            r.get("ssl_status"),
            r.get("error"),
        ])
    return rows


def export_excel(results: list, output_file: str, include_history: bool = True):
    wb = Workbook()
    wb.remove(wb.active)

    # ── Results sheet ──
    ws = wb.create_sheet("Audit Results")
    headers = [col[0] for col in COLUMNS]
    ws.append(headers)
    style_header(ws)

    for row in results_to_rows(results):
        ws.append(row)

    # Column widths
    for i, (_, width) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Color overall status column (col 3)
    apply_status_colors(ws, col_index=3)

    # Freeze top row
    ws.freeze_panes = "A2"

    # ── Summary sheet ──
    ws_sum = wb.create_sheet("Summary")
    ws_sum.append(["Metric", "Value"])
    style_header(ws_sum, COLOR_SUBHEADER)

    total = len(results)
    ok = sum(1 for r in results if r["overall_status"] == "OK")
    warnings = sum(1 for r in results if r["overall_status"] in ("WARNING", "SLOW"))
    errors = sum(1 for r in results if r["overall_status"] in ("ERROR", "DOWN"))
    avg_response = round(sum(r["response_time_ms"] or 0 for r in results) / max(total, 1))
    ssl_expiring = sum(1 for r in results if r.get("ssl_days_left") and r["ssl_days_left"] <= SSL_WARN_DAYS)

    summary_rows = [
        ("Sites Audited", total),
        ("✅ OK", ok),
        ("⚠️ Warnings / Slow", warnings),
        ("❌ Errors / Down", errors),
        ("Avg Response Time (ms)", avg_response),
        ("SSL Expiring Soon", ssl_expiring),
        ("Audit Time", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for row in summary_rows:
        ws_sum.append(row)

    ws_sum.column_dimensions["A"].width = 24
    ws_sum.column_dimensions["B"].width = 20

    # ── History sheet ──
    if include_history:
        history = load_history()
        if history:
            ws_hist = wb.create_sheet("History")
            ws_hist.append(headers)
            style_header(ws_hist, COLOR_SUBHEADER)
            for row in results_to_rows(history[-200:]):
                ws_hist.append(row)
            for i, (_, width) in enumerate(COLUMNS, 1):
                ws_hist.column_dimensions[get_column_letter(i)].width = width
            apply_status_colors(ws_hist, col_index=3)
            ws_hist.freeze_panes = "A2"

    output = output_file if output_file.endswith(".xlsx") else output_file + ".xlsx"
    wb.save(output)
    print(f"  ✅ Report saved: {output}")
    return output


# ─── SCHEDULER ────────────────────────────────────────────────────────────────

def run_audit(urls: list, output_file: str):
    """Single audit run — used both standalone and by scheduler."""
    print(f"\n🔍 Running audit — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    results = audit_all(urls)
    print_results(results)
    save_to_history(results)
    if output_file:
        export_excel(results, output_file)
    return results


def run_scheduled(urls: list, interval_minutes: int, output_file: str):
    """Run audit on a schedule indefinitely."""
    print(f"\n⏱️  Scheduled mode — running every {interval_minutes} minute(s). Ctrl+C to stop.\n")
    run_audit(urls, output_file)  # run immediately first
    schedule.every(interval_minutes).minutes.do(run_audit, urls=urls, output_file=output_file)
    try:
        while True:
            schedule.run_pending()
            time.sleep(10)
    except KeyboardInterrupt:
        print("\n\nMonitoring stopped.")


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Audit websites for performance, uptime, and SSL health.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python sitemon.py --urls https://example.com https://another.com
  python sitemon.py --file sites.txt
  python sitemon.py --file sites.txt --output report.xlsx
  python sitemon.py --file sites.txt --schedule 30
  python sitemon.py --history
        """
    )

    input_group = parser.add_mutually_exclusive_group()
    input_group.add_argument("--urls", nargs="+", help="One or more URLs to audit")
    input_group.add_argument("--file", type=str, help="Text file with one URL per line")

    parser.add_argument("--output", type=str, default="site_audit.xlsx", help="Output Excel filename")
    parser.add_argument("--schedule", type=int, metavar="MINUTES", help="Run on a schedule every N minutes")
    parser.add_argument("--history", action="store_true", help="Show recent audit history")
    parser.add_argument("--no-export", action="store_true", help="Skip Excel export")

    args = parser.parse_args()

    if args.history:
        show_history()
        return

    if not args.urls and not args.file:
        parser.print_help()
        sys.exit(1)

    print("\n🖥️  Site Performance Monitor")
    print("-" * 40)

    urls = args.urls if args.urls else load_urls_from_file(args.file)
    output = None if args.no_export else args.output

    if args.schedule:
        run_scheduled(urls, args.schedule, output)
    else:
        run_audit(urls, output)


if __name__ == "__main__":
    main()
